from rest_framework import viewsets, status
from rest_framework.decorators import action, api_view, permission_classes
from rest_framework.response import Response
from rest_framework import permissions
from django.db.models import Count, Sum, Avg, Q, F, DecimalField
from django.db.models.functions import TruncMonth, ExtractYear, ExtractMonth
from django.utils import timezone
from django.http import HttpResponse
from django.conf import settings
from django.template.loader import render_to_string
import csv
from datetime import datetime, timedelta
from decimal import Decimal
from collections import defaultdict
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from weasyprint import HTML

from students.models import Student
from payments.models import Payment, Scholarship
from academics.models import CourseEnrollment, Course, Career
from .serializers import (
    StudentReportSerializer, PaymentReportSerializer,
    AcademicReportSerializer, ScholarshipReportSerializer,
    OverviewReportSerializer
)
from users.permissions import HasPermission


class ReportsViewSet(viewsets.ViewSet):
    """ViewSet para generar reportes consolidados"""
    
    def get_permissions(self):
        """Permisos específicos por acción - más permisivo durante desarrollo"""
        # Temporalmente más permisivo - solo requiere autenticación
        # TODO: Restaurar permiso específico en producción
        # return [permissions.IsAuthenticated(), HasPermission('view_reports')]
        return [permissions.IsAuthenticated()]

    @action(detail=False, methods=['get'], url_path='overview')
    def overview(self, request):
        """
        Reporte general consolidado con todas las estadísticas
        Permite filtrar por rango de fechas
        """
        # Obtener parámetros de filtro
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        
        # Parsear fechas o usar valores por defecto
        if start_date:
            try:
                start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
            except ValueError:
                start_date = None
        else:
            # Por defecto, año actual
            start_date = datetime.now().replace(month=1, day=1).date()
        
        if end_date:
            try:
                end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
            except ValueError:
                end_date = None
        else:
            end_date = timezone.now().date()
        
        # Obtener reportes individuales
        student_data = self._get_student_report(start_date, end_date)
        payment_data = self._get_payment_report(start_date, end_date)
        academic_data = self._get_academic_report(start_date, end_date)
        scholarship_data = self._get_scholarship_report(start_date, end_date)
        
        # Consolidar respuesta
        report_data = {
            'students': student_data,
            'payments': payment_data,
            'academics': academic_data,
            'scholarships': scholarship_data,
            'date_range': {
                'start': start_date.isoformat() if start_date else None,
                'end': end_date.isoformat() if end_date else None,
            }
        }
        
        # Devolver datos directamente (el serializador es opcional para validación)
        # En este caso, los datos ya están estructurados correctamente
        return Response(report_data, status=status.HTTP_200_OK)

    @action(detail=False, methods=['get'], url_path='students')
    def students_report(self, request):
        """Reporte detallado de estudiantes"""
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        
        if start_date:
            try:
                start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
            except ValueError:
                start_date = None
        else:
            start_date = datetime.now().replace(month=1, day=1).date()
        
        if end_date:
            try:
                end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
            except ValueError:
                end_date = None
        else:
            end_date = timezone.now().date()
        
        data = self._get_student_report(start_date, end_date)
        serializer = StudentReportSerializer(data)
        return Response(serializer.data)

    @action(detail=False, methods=['get'], url_path='payments')
    def payments_report(self, request):
        """Reporte detallado de pagos"""
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        
        if start_date:
            try:
                start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
            except ValueError:
                start_date = None
        else:
            start_date = datetime.now().replace(month=1, day=1).date()
        
        if end_date:
            try:
                end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
            except ValueError:
                end_date = None
        else:
            end_date = timezone.now().date()
        
        data = self._get_payment_report(start_date, end_date)
        serializer = PaymentReportSerializer(data)
        return Response(serializer.data)

    @action(detail=False, methods=['get'], url_path='academics')
    def academics_report(self, request):
        """Reporte detallado académico"""
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        
        if start_date:
            try:
                start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
            except ValueError:
                start_date = None
        else:
            start_date = datetime.now().replace(month=1, day=1).date()
        
        if end_date:
            try:
                end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
            except ValueError:
                end_date = None
        else:
            end_date = timezone.now().date()
        
        data = self._get_academic_report(start_date, end_date)
        serializer = AcademicReportSerializer(data)
        return Response(serializer.data)

    @action(detail=False, methods=['get'], url_path='scholarships')
    def scholarships_report(self, request):
        """Reporte detallado de becas"""
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        
        if start_date:
            try:
                start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
            except ValueError:
                start_date = None
        else:
            start_date = datetime.now().replace(month=1, day=1).date()
        
        if end_date:
            try:
                end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
            except ValueError:
                end_date = None
        else:
            end_date = timezone.now().date()
        
        data = self._get_scholarship_report(start_date, end_date)
        serializer = ScholarshipReportSerializer(data)
        return Response(serializer.data)

    @action(detail=False, methods=['get'], url_path='export/students')
    def export_students_csv(self, request):
        """Exportar reporte de estudiantes a CSV"""
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        
        if start_date:
            try:
                start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
            except ValueError:
                start_date = None
        else:
            start_date = datetime.now().replace(month=1, day=1).date()
        
        if end_date:
            try:
                end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
            except ValueError:
                end_date = None
        else:
            end_date = timezone.now().date()
        
        students_qs = Student.objects.all()
        if start_date and end_date:
            students_qs = students_qs.filter(enrollment_date__gte=start_date, enrollment_date__lte=end_date)
        
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="estudiantes_{datetime.now().strftime("%Y%m%d")}.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['Carnet', 'Nombre Completo', 'Email', 'Carrera', 'Fecha Inscripción', 'Activo', 'Tipo Beca'])
        
        for student in students_qs:
            writer.writerow([
                student.carnet or '',
                student.get_full_name(),
                student.email,
                student.career.name if student.career else '',
                student.enrollment_date.strftime('%Y-%m-%d'),
                'Sí' if student.is_active else 'No',
                student.get_scholarship_type_display(),
            ])
        
        return response

    @action(detail=False, methods=['get'], url_path='export/payments')
    def export_payments_csv(self, request):
        """Exportar reporte de pagos a CSV"""
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        
        if start_date:
            try:
                start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
            except ValueError:
                start_date = None
        else:
            start_date = datetime.now().replace(month=1, day=1).date()
        
        if end_date:
            try:
                end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
            except ValueError:
                end_date = None
        else:
            end_date = timezone.now().date()
        
        payments_qs = Payment.objects.select_related('student', 'payment_type').all()
        if start_date and end_date:
            # Filtrar por payment_date si existe, o por created_at si payment_date es None
            payments_qs = payments_qs.filter(
                Q(payment_date__gte=start_date, payment_date__lte=end_date) |
                Q(payment_date__isnull=True, created_at__date__gte=start_date, created_at__date__lte=end_date)
            )
        
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="pagos_{datetime.now().strftime("%Y%m%d")}.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['Estudiante', 'Tipo Pago', 'Monto', 'Mes', 'Año', 'Estado', 'Fecha Pago', 'Método Pago'])
        
        for payment in payments_qs:
            writer.writerow([
                payment.student.get_full_name(),
                payment.payment_type.name if payment.payment_type else '',
                str(payment.amount),
                payment.get_month_display() if payment.month else '',
                str(payment.year) if payment.year else '',
                payment.get_status_display(),
                payment.payment_date.strftime('%Y-%m-%d'),
                payment.get_payment_method_display(),
            ])
        
        return response

    def _get_student_report(self, start_date, end_date):
        """Generar datos de reporte de estudiantes"""
        # Base queryset de estudiantes
        students_qs = Student.objects.all()
        
        # Filtrar por fecha de inscripción si se especifica (más relevante que created_at)
        if start_date and end_date:
            students_qs = students_qs.filter(enrollment_date__gte=start_date, enrollment_date__lte=end_date)
        
        total = students_qs.count()
        active = students_qs.filter(is_active=True).count()
        
        # Estudiantes por carrera
        students_by_career = {}
        career_data = students_qs.values('career__name').annotate(count=Count('id'))
        for item in career_data:
            career_name = item['career__name'] or 'Sin carrera'
            students_by_career[career_name] = item['count']
        
        # Estudiantes por tipo de beca
        students_by_scholarship = {}
        scholarship_data = students_qs.values('scholarship_type').annotate(count=Count('id'))
        for item in scholarship_data:
            scholarship_type = item['scholarship_type'] or 'NINGUNA'
            students_by_scholarship[scholarship_type] = item['count']
        
        return {
            'total': total,
            'active': active,
            'by_career': students_by_career,
            'by_scholarship': students_by_scholarship,
        }

    def _get_payment_report(self, start_date, end_date):
        """Generar datos de reporte de pagos"""
        payments_qs = Payment.objects.all()
        
        # Filtrar por rango de fechas
        if start_date and end_date:
            # Filtrar por payment_date si existe, o por created_at si payment_date es None
            payments_qs = payments_qs.filter(
                Q(payment_date__gte=start_date, payment_date__lte=end_date) |
                Q(payment_date__isnull=True, created_at__date__gte=start_date, created_at__date__lte=end_date)
            )
        
        total = payments_qs.count()
        approved = payments_qs.filter(status='APROBADO').count()
        pending = payments_qs.filter(status__in=['PENDIENTE', 'EN_REVISION']).count()
        rejected = payments_qs.filter(status='RECHAZADO').count()
        
        # Monto total de pagos aprobados
        total_amount = payments_qs.filter(status='APROBADO').aggregate(
            total=Sum('amount', output_field=DecimalField(max_digits=12, decimal_places=2))
        )['total'] or Decimal('0.00')
        
        # Pagos por mes
        payments_by_month = {}
        approved_payments = payments_qs.filter(status='APROBADO')
        
        if start_date and end_date:
            # Agrupar por año-mes usando payment_date si existe, sino usar created_at
            # Primero pagos con payment_date
            monthly_data_with_date = approved_payments.filter(
                payment_date__gte=start_date,
                payment_date__lte=end_date
            ).exclude(payment_date__isnull=True).annotate(
                year_extracted=ExtractYear('payment_date'),
                month_extracted=ExtractMonth('payment_date')
            ).values('year_extracted', 'month_extracted').annotate(
                total_amount=Sum('amount', output_field=DecimalField(max_digits=12, decimal_places=2))
            ).order_by('year_extracted', 'month_extracted')
            
            # Pagos sin payment_date usando created_at
            monthly_data_no_date = approved_payments.filter(
                payment_date__isnull=True,
                created_at__date__gte=start_date,
                created_at__date__lte=end_date
            ).annotate(
                year_extracted=ExtractYear('created_at'),
                month_extracted=ExtractMonth('created_at')
            ).values('year_extracted', 'month_extracted').annotate(
                total_amount=Sum('amount', output_field=DecimalField(max_digits=12, decimal_places=2))
            ).order_by('year_extracted', 'month_extracted')
            
            # Combinar resultados
            combined_monthly = defaultdict(lambda: {'year_extracted': None, 'month_extracted': None, 'total_amount': Decimal('0.00')})
            for item in monthly_data_with_date:
                key = (item['year_extracted'], item['month_extracted'])
                combined_monthly[key]['year_extracted'] = item['year_extracted']
                combined_monthly[key]['month_extracted'] = item['month_extracted']
                combined_monthly[key]['total_amount'] += (item['total_amount'] or Decimal('0.00'))
            for item in monthly_data_no_date:
                key = (item['year_extracted'], item['month_extracted'])
                if key in combined_monthly:
                    combined_monthly[key]['total_amount'] += (item['total_amount'] or Decimal('0.00'))
                else:
                    combined_monthly[key] = item
            monthly_data = list(combined_monthly.values())
            
            for item in monthly_data:
                month_key = f"{item['year_extracted']}-{str(item['month_extracted']).zfill(2)}"
                payments_by_month[month_key] = float(item['total_amount'] or Decimal('0.00'))
        else:
            # Usar mes y año del modelo si están disponibles
            monthly_data = approved_payments.exclude(year__isnull=True, month__isnull=True).values(
                'year', 'month'
            ).annotate(
                total_amount=Sum('amount', output_field=DecimalField(max_digits=12, decimal_places=2))
            ).order_by('year', 'month')
            
            for item in monthly_data:
                month_key = f"{item['year']}-{str(item['month']).zfill(2)}"
                payments_by_month[month_key] = float(item['total_amount'] or Decimal('0.00'))
        
        return {
            'total': total,
            'approved': approved,
            'pending': pending,
            'rejected': rejected,
            'total_amount': float(total_amount),
            'by_month': payments_by_month,
        }

    def _get_academic_report(self, start_date, end_date):
        """Generar datos de reporte académico"""
        enrollments_qs = CourseEnrollment.objects.all()
        
        # Filtrar por rango de fechas
        if start_date and end_date:
            enrollments_qs = enrollments_qs.filter(
                Q(enrollment_date__gte=start_date) & Q(enrollment_date__lte=end_date)
            )
        
        total_enrollments = enrollments_qs.count()
        approved_courses = enrollments_qs.filter(status='APROBADO').count()
        
        # Promedio de calificaciones
        avg_data = enrollments_qs.filter(
            status='APROBADO',
            final_grade__isnull=False
        ).aggregate(avg_grade=Avg('final_grade'))
        
        average_grade = float(avg_data['avg_grade'] or Decimal('0.00'))
        
        # Completitud del pensum (cursos aprobados / total de cursos por carrera)
        # Calcular el total de cursos por carrera
        total_students = Student.objects.filter(is_active=True).count()
        total_courses_by_career = {}
        
        careers = Career.objects.all()
        for career in careers:
            total_courses = Course.objects.filter(career=career).count()
            if total_courses > 0:
                total_courses_by_career[career.id] = total_courses
        
        # Calcular completitud
        if total_students > 0 and total_courses_by_career:
            total_possible_courses = sum(total_courses_by_career.values()) * total_students
            if total_possible_courses > 0:
                pensum_completion = (approved_courses / total_possible_courses) * 100
            else:
                pensum_completion = Decimal('0.00')
        else:
            pensum_completion = Decimal('0.00')
        
        pensum_completion = min(float(pensum_completion), 100.0)
        
        return {
            'total_enrollments': total_enrollments,
            'approved_courses': approved_courses,
            'average_grade': average_grade,
            'pensum_completion': pensum_completion,
        }

    def _get_scholarship_report(self, start_date, end_date):
        """Generar datos de reporte de becas"""
        scholarships_qs = Scholarship.objects.all()
        
        # Filtrar por fecha de inicio si se especifica
        if start_date and end_date:
            scholarships_qs = scholarships_qs.filter(
                Q(start_date__gte=start_date) & Q(start_date__lte=end_date)
            )
        
        total = scholarships_qs.count()
        active = scholarships_qs.filter(status='ACTIVA').count()
        completa = scholarships_qs.filter(
            status='ACTIVA',
            scholarship_type='COMPLETA'
        ).count()
        media = scholarships_qs.filter(
            status='ACTIVA',
            scholarship_type='MEDIA'
        ).count()
        
        return {
            'total': total,
            'active': active,
            'completa': completa,
            'media': media,
        }

    @action(detail=False, methods=['get'], url_path='export/payments/excel')
    def export_payments_excel(self, request):
        """Exportar reporte detallado de pagos a Excel"""
        try:
            start_date = request.query_params.get('start_date')
            end_date = request.query_params.get('end_date')
            
            # Parsear fechas
            if start_date:
                try:
                    start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
                except ValueError:
                    start_date = None
            else:
                start_date = (datetime.now() - timedelta(days=365)).date()
            
            if end_date:
                try:
                    end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
                except ValueError:
                    end_date = None
            else:
                end_date = timezone.now().date()
            
            # Obtener pagos con relaciones
            payments_qs = Payment.objects.select_related(
                'student', 
                'payment_type',
                'created_by',
                'approved_by'
            ).order_by('-payment_date', '-created_at')
            
            # Filtrar por rango de fechas
            if start_date and end_date:
                payments_qs = payments_qs.filter(
                    Q(payment_date__gte=start_date, payment_date__lte=end_date) |
                    Q(payment_date__isnull=True, created_at__date__gte=start_date, created_at__date__lte=end_date)
                )
            
            # Crear workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Reporte de Pagos"
            
            # Estilos
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            title_font = Font(bold=True, size=14)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Título
            ws.merge_cells('A1:Q1')
            title_cell = ws['A1']
            title_cell.value = f"REPORTE DETALLADO DE PAGOS - {start_date} a {end_date}"
            title_cell.font = title_font
            title_cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Encabezados
            headers = [
                'Fecha Pago', 'Fecha Creación', 'Carnet', 'Estudiante', 'Email',
                'Tipo de Pago', 'Monto Original', 'Descuento Beca', 'Mora', 'Monto Final',
                'Estado', 'Método de Pago', 'Referencia', 'Mes', 'Año',
                'Creado Por', 'Aprobado Por', 'Fecha Aprobación'
            ]
            
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col_num)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = border
                ws.column_dimensions[get_column_letter(col_num)].width = 18
            
            # Datos
            row_num = 4
            for payment in payments_qs:
                original_amount = payment.base_amount if payment.base_amount else payment.amount
                final_amount = payment.amount
                descuento_beca = getattr(payment, 'scholarship_discount_amount', None) or Decimal('0.00')
                
                row_data = [
                    payment.payment_date.strftime('%Y-%m-%d') if payment.payment_date else '',
                    payment.created_at.strftime('%Y-%m-%d %H:%M:%S') if payment.created_at else '',
                    payment.student.carnet or '',
                    payment.student.get_full_name(),
                    payment.student.email or '',
                    payment.payment_type.name if payment.payment_type else 'Sin tipo',
                    float(original_amount or Decimal('0.00')),
                    float(descuento_beca),
                    float(payment.penalty_amount or Decimal('0.00')),
                    float(final_amount or Decimal('0.00')),
                    payment.get_status_display(),
                    payment.get_payment_method_display(),
                    payment.payment_reference or '',
                    payment.month or '',
                    payment.year or '',
                    payment.created_by.get_full_name() if payment.created_by else '',
                    payment.approved_by.get_full_name() if payment.approved_by else '',
                    payment.approved_at.strftime('%Y-%m-%d %H:%M:%S') if payment.approved_at else '',
                ]
                
                for col_num, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.value = value
                    cell.border = border
                    if col_num in [7, 8, 9, 10]:  # Columnas de montos
                        cell.number_format = '#,##0.00'
                        cell.alignment = Alignment(horizontal='right')
                    else:
                        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                
                row_num += 1
            
            # Resumen
            total_pagos = payments_qs.count()
            total_aprobados = payments_qs.filter(status='APROBADO').count()
            monto_total = payments_qs.filter(status='APROBADO').aggregate(
                total=Sum('amount', output_field=DecimalField(max_digits=12, decimal_places=2))
            )['total'] or Decimal('0.00')
            
            ws.merge_cells(f'A{row_num+1}:Q{row_num+1}')
            summary_cell = ws.cell(row=row_num+1, column=1)
            summary_cell.value = f"RESUMEN: Total Pagos: {total_pagos} | Aprobados: {total_aprobados} | Monto Total: ${float(monto_total):,.2f}"
            summary_cell.font = Font(bold=True, size=11)
            summary_cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            summary_cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Preparar respuesta
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            filename = f"reporte_pagos_detallado_{start_date}_{end_date}.xlsx"
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            
            wb.save(response)
            return response
            
        except Exception as e:
            import traceback
            print(f"Error en export_payments_excel: {str(e)}")
            print(traceback.format_exc())
            return Response({
                'error': 'Error al generar el archivo Excel',
                'detail': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['get'], url_path='payments/detailed')
    def payments_detailed(self, request):
        """
        Reporte detallado de todos los pagos recibidos
        Incluye: fecha, estudiante, tipo de pago, monto, estado, método de pago
        """
        try:
            start_date = request.query_params.get('start_date')
            end_date = request.query_params.get('end_date')
            
            # Parsear fechas
            if start_date:
                try:
                    start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
                except ValueError:
                    start_date = None
            else:
                # Por defecto, último año
                start_date = (datetime.now() - timedelta(days=365)).date()
            
            if end_date:
                try:
                    end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
                except ValueError:
                    end_date = None
            else:
                end_date = timezone.now().date()
            
            # Obtener pagos con relaciones
            payments_qs = Payment.objects.select_related(
                'student', 
                'payment_type',
                'created_by',
                'approved_by'
            ).order_by('-payment_date', '-created_at')
            
            # Filtrar por rango de fechas
            if start_date and end_date:
                # Filtrar por payment_date si existe, o por created_at si payment_date es None
                payments_qs = payments_qs.filter(
                    Q(payment_date__gte=start_date, payment_date__lte=end_date) |
                    Q(payment_date__isnull=True, created_at__date__gte=start_date, created_at__date__lte=end_date)
                )
            
            # Construir lista de pagos detallados
            payments_list = []
            for payment in payments_qs:
                try:
                    # Calcular montos: usar base_amount si existe, sino usar amount
                    original_amount = payment.base_amount if payment.base_amount else payment.amount
                    final_amount = payment.amount  # amount ya incluye mora si existe
                    descuento_beca = getattr(payment, 'scholarship_discount_amount', None) or Decimal('0.00')
                    
                    # Validar que el estudiante exista
                    if not payment.student:
                        continue  # Saltar pagos sin estudiante
                    
                    payments_list.append({
                        'id': str(payment.id),
                        'fecha_pago': payment.payment_date.strftime('%Y-%m-%d') if payment.payment_date else None,
                        'fecha_creacion': payment.created_at.strftime('%Y-%m-%d %H:%M:%S') if payment.created_at else None,
                        'estudiante': {
                            'id': str(payment.student.id),
                            'carnet': payment.student.carnet or '',
                            'nombre_completo': payment.student.get_full_name() if hasattr(payment.student, 'get_full_name') else f"{payment.student.first_name or ''} {payment.student.first_last_name or ''}".strip(),
                            'email': payment.student.email or '',
                        },
                        'tipo_pago': {
                            'codigo': payment.payment_type.code if payment.payment_type else '',
                            'nombre': payment.payment_type.name if payment.payment_type else 'Sin tipo',
                        },
                        'monto': {
                            'original': float(original_amount or Decimal('0.00')),
                            'descuento_beca': float(descuento_beca),
                            'mora': float(payment.penalty_amount or Decimal('0.00')),
                            'final': float(final_amount or Decimal('0.00')),
                        },
                        'estado': payment.status,
                        'estado_display': payment.get_status_display(),
                        'metodo_pago': payment.payment_method,
                        'metodo_pago_display': payment.get_payment_method_display(),
                        'referencia': payment.payment_reference or '',
                        'mes': payment.month,
                        'año': payment.year,
                        'creado_por': payment.created_by.get_full_name() if payment.created_by and hasattr(payment.created_by, 'get_full_name') else (f"{payment.created_by.first_name or ''} {payment.created_by.last_name or ''}".strip() if payment.created_by else None),
                        'aprobado_por': payment.approved_by.get_full_name() if payment.approved_by and hasattr(payment.approved_by, 'get_full_name') else (f"{payment.approved_by.first_name or ''} {payment.approved_by.last_name or ''}".strip() if payment.approved_by else None),
                        'fecha_aprobacion': payment.approved_at.strftime('%Y-%m-%d %H:%M:%S') if payment.approved_at else None,
                    })
                except Exception as e:
                    # Log del error pero continuar con el siguiente pago
                    import traceback
                    print(f"Error procesando pago {payment.id}: {str(e)}")
                    print(traceback.format_exc())
                    continue
            
            # Estadísticas resumidas
            total_pagos = payments_qs.count()
            total_aprobados = payments_qs.filter(status='APROBADO').count()
            total_pendientes = payments_qs.filter(status__in=['PENDIENTE', 'EN_REVISION']).count()
            total_rechazados = payments_qs.filter(status='RECHAZADO').count()
            
            # Usar amount directamente ya que final_amount puede no existir
            monto_total_aprobados = payments_qs.filter(status='APROBADO').aggregate(
                total=Sum('amount', output_field=DecimalField(max_digits=12, decimal_places=2))
            )['total'] or Decimal('0.00')
            
            return Response({
                'resumen': {
                    'total_pagos': total_pagos,
                    'aprobados': total_aprobados,
                    'pendientes': total_pendientes,
                    'rechazados': total_rechazados,
                    'monto_total_aprobados': float(monto_total_aprobados),
                    'rango_fechas': {
                        'inicio': start_date.isoformat() if start_date else None,
                        'fin': end_date.isoformat() if end_date else None,
                    }
                },
                'pagos': payments_list,
            }, status=status.HTTP_200_OK)
        except Exception as e:
            import traceback
            error_detail = str(e)
            traceback_str = traceback.format_exc()
            print(f"Error en payments_detailed: {error_detail}")
            print(traceback_str)
            return Response({
                'error': 'Error al generar el reporte detallado de pagos',
                'detail': error_detail
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['get'], url_path='executive')
    def executive_report(self, request):
        """
        Reporte ejecutivo para gerencia
        Incluye estadísticas agrupadas por tipo de pago, métodos de pago, etc.
        """
        try:
            start_date = request.query_params.get('start_date')
            end_date = request.query_params.get('end_date')
            
            # Parsear fechas
            if start_date:
                try:
                    start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
                except ValueError:
                    start_date = None
            else:
                # Por defecto, año actual
                start_date = datetime.now().replace(month=1, day=1).date()
            
            if end_date:
                try:
                    end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
                except ValueError:
                    end_date = None
            else:
                end_date = timezone.now().date()
            
            # Obtener pagos aprobados (solo estos cuentan para estadísticas financieras)
            payments_qs = Payment.objects.select_related(
                'student', 
                'payment_type'
            ).filter(status='APROBADO')
            
            # Filtrar por rango de fechas
            if start_date and end_date:
                # Filtrar por payment_date si existe, o por created_at si payment_date es None
                payments_qs = payments_qs.filter(
                    Q(payment_date__gte=start_date, payment_date__lte=end_date) |
                    Q(payment_date__isnull=True, created_at__date__gte=start_date, created_at__date__lte=end_date)
                )
            
            # 1. Estadísticas generales
            total_pagos = payments_qs.count()
            monto_total = payments_qs.aggregate(
                total=Sum('amount', output_field=DecimalField(max_digits=12, decimal_places=2))
            )['total'] or Decimal('0.00')
            
            promedio_pago = monto_total / total_pagos if total_pagos > 0 else Decimal('0.00')
            
            # 2. Pagos agrupados por tipo de pago
            pagos_por_tipo = {}
            pagos_por_tipo_data = payments_qs.values(
                'payment_type__code',
                'payment_type__name'
            ).annotate(
                cantidad=Count('id'),
                monto_total=Sum('amount', output_field=DecimalField(max_digits=12, decimal_places=2))
            ).order_by('-monto_total')
            
            for item in pagos_por_tipo_data:
                tipo_key = item['payment_type__code'] or 'SIN_TIPO'
                pagos_por_tipo[tipo_key] = {
                    'codigo': item['payment_type__code'] or '',
                    'nombre': item['payment_type__name'] or 'Sin tipo',
                    'cantidad': item['cantidad'],
                    'monto_total': float(item['monto_total'] or Decimal('0.00')),
                    'porcentaje_del_total': float((item['monto_total'] / monto_total * 100) if monto_total > 0 else Decimal('0.00')),
                }
            
            # 3. Pagos agrupados por método de pago
            pagos_por_metodo = {}
            pagos_por_metodo_data = payments_qs.values('payment_method').annotate(
                cantidad=Count('id'),
                monto_total=Sum('amount', output_field=DecimalField(max_digits=12, decimal_places=2))
            ).order_by('-monto_total')
            
            for item in pagos_por_metodo_data:
                metodo_key = item['payment_method'] or 'SIN_METODO'
                pagos_por_metodo[metodo_key] = {
                    'metodo': item['payment_method'],
                    'metodo_display': dict(Payment.PAYMENT_METHODS).get(item['payment_method'], item['payment_method']),
                    'cantidad': item['cantidad'],
                    'monto_total': float(item['monto_total'] or Decimal('0.00')),
                    'porcentaje_del_total': float((item['monto_total'] / monto_total * 100) if monto_total > 0 else Decimal('0.00')),
                }
            
            # 4. Pagos por mes (últimos 12 meses o rango especificado)
            pagos_por_mes = {}
            # Usar payment_date si existe, sino usar created_at
            monthly_data_with_date = payments_qs.exclude(payment_date__isnull=True).annotate(
                year_extracted=ExtractYear('payment_date'),
                month_extracted=ExtractMonth('payment_date')
            ).values('year_extracted', 'month_extracted').annotate(
                cantidad=Count('id'),
                monto_total=Sum('amount', output_field=DecimalField(max_digits=12, decimal_places=2))
            ).order_by('year_extracted', 'month_extracted')
            
            # Pagos sin payment_date usando created_at
            monthly_data_no_date = payments_qs.filter(payment_date__isnull=True).annotate(
                year_extracted=ExtractYear('created_at'),
                month_extracted=ExtractMonth('created_at')
            ).values('year_extracted', 'month_extracted').annotate(
                cantidad=Count('id'),
                monto_total=Sum('amount', output_field=DecimalField(max_digits=12, decimal_places=2))
            ).order_by('year_extracted', 'month_extracted')
            
            # Combinar resultados
            combined_monthly = defaultdict(lambda: {'year_extracted': None, 'month_extracted': None, 'cantidad': 0, 'monto_total': Decimal('0.00')})
            
            for item in monthly_data_with_date:
                if item['year_extracted'] is not None and item['month_extracted'] is not None:
                    key = (item['year_extracted'], item['month_extracted'])
                    combined_monthly[key]['year_extracted'] = item['year_extracted']
                    combined_monthly[key]['month_extracted'] = item['month_extracted']
                    combined_monthly[key]['cantidad'] += item['cantidad']
                    combined_monthly[key]['monto_total'] += (item['monto_total'] or Decimal('0.00'))
            
            for item in monthly_data_no_date:
                if item['year_extracted'] is not None and item['month_extracted'] is not None:
                    key = (item['year_extracted'], item['month_extracted'])
                    if key in combined_monthly:
                        combined_monthly[key]['cantidad'] += item['cantidad']
                        combined_monthly[key]['monto_total'] += (item['monto_total'] or Decimal('0.00'))
                    else:
                        combined_monthly[key]['year_extracted'] = item['year_extracted']
                        combined_monthly[key]['month_extracted'] = item['month_extracted']
                        combined_monthly[key]['cantidad'] = item['cantidad']
                        combined_monthly[key]['monto_total'] = (item['monto_total'] or Decimal('0.00'))
            
            monthly_data = [item for item in combined_monthly.values() if item['year_extracted'] is not None and item['month_extracted'] is not None]
            
            meses_es = {
                1: 'Enero', 2: 'Febrero', 3: 'Marzo', 4: 'Abril',
                5: 'Mayo', 6: 'Junio', 7: 'Julio', 8: 'Agosto',
                9: 'Septiembre', 10: 'Octubre', 11: 'Noviembre', 12: 'Diciembre'
            }
            
            for item in monthly_data:
                if item['year_extracted'] is not None and item['month_extracted'] is not None:
                    month_key = f"{item['year_extracted']}-{str(item['month_extracted']).zfill(2)}"
                    pagos_por_mes[month_key] = {
                        'año': item['year_extracted'],
                        'mes': item['month_extracted'],
                        'mes_nombre': meses_es.get(item['month_extracted'], ''),
                        'cantidad': item['cantidad'],
                        'monto_total': float(item['monto_total'] or Decimal('0.00')),
                    }
            
            # 5. Top estudiantes por monto pagado
            top_estudiantes = payments_qs.values(
                'student__id',
                'student__carnet',
                'student__first_name',
                'student__first_last_name',
                'student__second_last_name'
            ).annotate(
                cantidad_pagos=Count('id'),
                monto_total=Sum('amount', output_field=DecimalField(max_digits=12, decimal_places=2))
            ).order_by('-monto_total')[:10]
            
            top_estudiantes_list = []
            for item in top_estudiantes:
                nombre_completo = f"{item['student__first_name'] or ''} {item['student__first_last_name'] or ''} {item['student__second_last_name'] or ''}".strip()
                top_estudiantes_list.append({
                    'carnet': item['student__carnet'] or '',
                    'nombre_completo': nombre_completo,
                    'cantidad_pagos': item['cantidad_pagos'],
                    'monto_total': float(item['monto_total'] or Decimal('0.00')),
                })
            
            return Response({
                'resumen_general': {
                    'total_pagos': total_pagos,
                    'monto_total': float(monto_total),
                    'promedio_por_pago': float(promedio_pago),
                    'rango_fechas': {
                        'inicio': start_date.isoformat() if start_date else None,
                        'fin': end_date.isoformat() if end_date else None,
                    }
                },
                'pagos_por_tipo': pagos_por_tipo,
                'pagos_por_metodo': pagos_por_metodo,
                'pagos_por_mes': pagos_por_mes,
                'top_estudiantes': top_estudiantes_list,
            }, status=status.HTTP_200_OK)
        except Exception as e:
            import traceback
            error_detail = str(e)
            traceback_str = traceback.format_exc()
            print(f"Error en executive_report: {error_detail}")
            print(traceback_str)
            return Response({
                'error': 'Error al generar el reporte ejecutivo',
                'detail': error_detail
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['get'], url_path='export/executive/pdf')
    def export_executive_pdf(self, request):
        """Exportar reporte ejecutivo a PDF con interpretación de datos"""
        try:
            start_date = request.query_params.get('start_date')
            end_date = request.query_params.get('end_date')
            
            # Parsear fechas
            if start_date:
                try:
                    start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
                except ValueError:
                    start_date = None
            else:
                start_date = datetime.now().replace(month=1, day=1).date()
            
            if end_date:
                try:
                    end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
                except ValueError:
                    end_date = None
            else:
                end_date = timezone.now().date()
            
            # Obtener datos del reporte ejecutivo (reutilizar lógica del método executive_report)
            payments_qs = Payment.objects.select_related(
                'student', 
                'payment_type'
            ).filter(status='APROBADO')
            
            if start_date and end_date:
                payments_qs = payments_qs.filter(
                    Q(payment_date__gte=start_date, payment_date__lte=end_date) |
                    Q(payment_date__isnull=True, created_at__date__gte=start_date, created_at__date__lte=end_date)
                )
            
            # Calcular estadísticas
            total_pagos = payments_qs.count()
            monto_total = payments_qs.aggregate(
                total=Sum('amount', output_field=DecimalField(max_digits=12, decimal_places=2))
            )['total'] or Decimal('0.00')
            promedio_pago = monto_total / total_pagos if total_pagos > 0 else Decimal('0.00')
            
            # Pagos por tipo
            pagos_por_tipo = {}
            pagos_por_tipo_data = payments_qs.values(
                'payment_type__code',
                'payment_type__name'
            ).annotate(
                cantidad=Count('id'),
                monto_total=Sum('amount', output_field=DecimalField(max_digits=12, decimal_places=2))
            ).order_by('-monto_total')
            
            for item in pagos_por_tipo_data:
                tipo_key = item['payment_type__code'] or 'SIN_TIPO'
                pagos_por_tipo[tipo_key] = {
                    'codigo': item['payment_type__code'] or '',
                    'nombre': item['payment_type__name'] or 'Sin tipo',
                    'cantidad': item['cantidad'],
                    'monto_total': float(item['monto_total'] or Decimal('0.00')),
                    'porcentaje_del_total': float((item['monto_total'] / monto_total * 100) if monto_total > 0 else Decimal('0.00')),
                }
            
            # Pagos por método
            pagos_por_metodo = {}
            pagos_por_metodo_data = payments_qs.values('payment_method').annotate(
                cantidad=Count('id'),
                monto_total=Sum('amount', output_field=DecimalField(max_digits=12, decimal_places=2))
            ).order_by('-monto_total')
            
            for item in pagos_por_metodo_data:
                metodo_key = item['payment_method'] or 'SIN_METODO'
                pagos_por_metodo[metodo_key] = {
                    'metodo': item['payment_method'],
                    'metodo_display': dict(Payment.PAYMENT_METHODS).get(item['payment_method'], item['payment_method']),
                    'cantidad': item['cantidad'],
                    'monto_total': float(item['monto_total'] or Decimal('0.00')),
                    'porcentaje_del_total': float((item['monto_total'] / monto_total * 100) if monto_total > 0 else Decimal('0.00')),
                }
            
            # Generar interpretación
            interpretacion = []
            if total_pagos > 0:
                interpretacion.append(f"Durante el período analizado ({start_date} a {end_date}), se registraron {total_pagos} pagos aprobados.")
                interpretacion.append(f"El monto total recaudado fue de ${float(monto_total):,.2f}, con un promedio de ${float(promedio_pago):,.2f} por pago.")
                
                if pagos_por_tipo:
                    tipo_mayor = max(pagos_por_tipo.values(), key=lambda x: x['monto_total'])
                    interpretacion.append(f"El tipo de pago más significativo fue '{tipo_mayor['nombre']}' con {tipo_mayor['cantidad']} pagos ({tipo_mayor['porcentaje_del_total']:.1f}% del total).")
                
                if pagos_por_metodo:
                    metodo_mayor = max(pagos_por_metodo.values(), key=lambda x: x['monto_total'])
                    interpretacion.append(f"El método de pago más utilizado fue '{metodo_mayor['metodo_display']}' con {metodo_mayor['cantidad']} transacciones ({metodo_mayor['porcentaje_del_total']:.1f}% del total).")
            
            # Renderizar HTML
            html_content = render_to_string('reports/executive_report_pdf.html', {
                'start_date': start_date,
                'end_date': end_date,
                'total_pagos': total_pagos,
                'monto_total': float(monto_total),
                'promedio_pago': float(promedio_pago),
                'pagos_por_tipo': pagos_por_tipo,
                'pagos_por_metodo': pagos_por_metodo,
                'interpretacion': interpretacion,
                'fecha_generacion': timezone.now().strftime('%d/%m/%Y %H:%M:%S'),
            })
            
            # Generar PDF
            html = HTML(string=html_content)
            pdf_file = html.write_pdf()
            
            response = HttpResponse(pdf_file, content_type='application/pdf')
            filename = f"reporte_ejecutivo_{start_date}_{end_date}.pdf"
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            
            return response
            
        except Exception as e:
            import traceback
            print(f"Error en export_executive_pdf: {str(e)}")
            print(traceback.format_exc())
            return Response({
                'error': 'Error al generar el archivo PDF',
                'detail': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
